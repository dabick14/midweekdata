#!/usr/bin/env python3
"""Update Midweek_Data.xlsx from Synago GraphQL stream governorship data."""

from __future__ import annotations

import argparse
import os
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

import requests
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

LOGIN_URL = "https://ndx3y4sa3znyoxzin6bzmoy6fi0jvruc.lambda-url.eu-west-2.on.aws/auth/login"
GRAPHQL_URL = "https://api-synago.firstlovecenter.com/graphql"

DEFAULT_STREAM_1 = "2dd77486-5d8d-4231-96e9-6d042500198a"
DEFAULT_STREAM_2 = "2d0f5804-0462-442f-93cc-25db95912589"
DEFAULT_STREAM_3 = "804e3aaf-e868-4772-a9f6-f0de76941d01"

COLUMN_HEADERS = [
    "Governorship",
    "Governor",
    "No. Of Bacentas",
    "Att",
    "Income(GHS)",
    "No. Of Services",
    "Services/Bacentas",
    "Services Not Held",
    "Comment",
]

GRAPHQL_QUERY = """
query getStreamGovernorships($id: ID!) {
  streams(where: {id: $id}) {
    id
    name
    leader {
      id
      firstName
      lastName
      fullName
      __typename
    }
    governorships {
      council {
        name
        leader {
          fullName
        }
      }
            bacentas {
                aggregateServiceRecords(limit: 1, skip: 0) {
                    id
                    attendance
                    income
                    numberOfServices
                    week
                    __typename
                }
            }
      name
      id
      stream_name
      bacentaCount
      aggregateServiceRecords(limit: 1, skip: 0) {
        id
        attendance
        income
        numberOfServices
        week
        __typename
      }
      services(limit: 1, skip: 0) {
        id
        createdAt
        attendance
        income
        week
        serviceDate {
          date
          __typename
        }
      }
      leader {
        id
        fullName
        __typename
      }
      __typename
    }
    __typename
  }
}
""".strip()


def _safe_int(value: Any, default: int = 0) -> int:
    if value is None:
        return default
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return default


def _normalize(text: Any) -> str:
    return str(text or "").strip().lower()


def sanitize_sheet_title(council_name: str, used_titles: set[str]) -> str:
    invalid_chars = set('[]:*?/\\')
    cleaned = "".join("-" if ch in invalid_chars else ch for ch in council_name).strip() or "Sheet"
    cleaned = cleaned[:31]

    if cleaned not in used_titles:
        used_titles.add(cleaned)
        return cleaned

    suffix = 2
    while True:
        suffix_text = f" ({suffix})"
        base = cleaned[: 31 - len(suffix_text)].rstrip()
        candidate = f"{base}{suffix_text}"
        if candidate not in used_titles:
            used_titles.add(candidate)
            return candidate
        suffix += 1


def get_required_env(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def load_dotenv(dotenv_path: Path) -> None:
    if not dotenv_path.exists():
        return

    for raw_line in dotenv_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue

        if line.startswith("export "):
            line = line[len("export ") :].strip()

        if "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()

        if len(value) >= 2 and value[0] == value[-1] and value[0] in {"\"", "'"}:
            value = value[1:-1]

        if key:
            os.environ.setdefault(key, value)


def login_get_access_token(email: str, password: str) -> str:
    payload = {"email": email, "password": password}
    headers = {"Content-Type": "application/json"}

    response = requests.post(LOGIN_URL, json=payload, headers=headers, timeout=30)
    response.raise_for_status()

    body = response.json()
    token = body.get("tokens", {}).get("accessToken")
    if not token:
        raise RuntimeError("Login succeeded but access token was missing in response")

    return token


def fetch_governorships_for_stream(stream_id: str, access_token: str) -> List[Dict[str, Any]]:
    payload = {"query": GRAPHQL_QUERY, "variables": {"id": stream_id}}
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {access_token}",
    }

    try:
        response = requests.post(GRAPHQL_URL, json=payload, headers=headers, timeout=45)
        response.raise_for_status()
    except requests.RequestException as exc:
        print(f"[WARN] Stream {stream_id}: request failed: {exc}")
        return []

    body = response.json()
    if body.get("errors"):
        print(f"[WARN] Stream {stream_id}: GraphQL returned errors: {body['errors']}")
        return []

    streams = body.get("data", {}).get("streams") or []
    if not streams:
        print(f"[WARN] Stream {stream_id}: no stream data returned")
        return []

    governorships = streams[0].get("governorships") or []
    return governorships


def resolve_metrics(governorship: Dict[str, Any], current_week: int) -> Tuple[int, int, int, str]:
    aggregate = (governorship.get("aggregateServiceRecords") or [None])[0]
    if aggregate and _safe_int(aggregate.get("week"), -1) == current_week:
        return (
            _safe_int(aggregate.get("attendance")),
            _safe_int(aggregate.get("income")),
            _safe_int(aggregate.get("numberOfServices")),
            "",
        )

    service = (governorship.get("services") or [None])[0]
    if service and _safe_int(service.get("week"), -1) == current_week:
        return (
            _safe_int(service.get("attendance")),
            _safe_int(service.get("income")),
            1,
            "Joint Service",
        )

    bacenta_matches = 0
    bacenta_attendance_total = 0
    bacenta_income_total = 0

    for bacenta in governorship.get("bacentas") or []:
        bacenta_aggregate = (bacenta.get("aggregateServiceRecords") or [None])[0]
        if not bacenta_aggregate:
            continue
        if _safe_int(bacenta_aggregate.get("week"), -1) != current_week:
            continue

        bacenta_matches += 1
        bacenta_attendance_total += _safe_int(bacenta_aggregate.get("attendance"))
        bacenta_income_total += _safe_int(bacenta_aggregate.get("income"))

    if bacenta_matches > 0:
        return (
            bacenta_attendance_total,
            bacenta_income_total,
            bacenta_matches,
            "Bacenta Sum",
        )

    return (0, 0, 0, "")


def transform_governorship(governorship: Dict[str, Any], current_week: int) -> Dict[str, Any]:
    name = str(governorship.get("name") or "").strip()
    governor_name = str((governorship.get("leader") or {}).get("fullName") or "").strip()
    bacenta_count = _safe_int(governorship.get("bacentaCount"))

    attendance, income_ghs, number_of_services, comment = resolve_metrics(governorship, current_week)
    services_not_held = max(bacenta_count - number_of_services, 0)

    return {
        "governorship": name,
        "governor": governor_name,
        "bacenta_count": bacenta_count,
        "attendance": attendance,
        "income_ghs": income_ghs,
        "number_of_services": number_of_services,
        "services_by_bacentas": f"{number_of_services}/{bacenta_count}",
        "services_not_held": services_not_held,
        "comment": comment,
    }


def collect_rows_by_sheet(
    stream_ids: Iterable[str], access_token: str, current_week: int
) -> Tuple[Dict[str, List[Dict[str, Any]]], Dict[str, str]]:
    rows_by_sheet: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    council_leaders: Dict[str, str] = {}

    for stream_id in stream_ids:
        governorships = fetch_governorships_for_stream(stream_id, access_token)
        print(f"[INFO] Stream {stream_id}: fetched {len(governorships)} governorship(s)")

        for governorship in governorships:
            council_name = str((governorship.get("council") or {}).get("name") or "").strip()
            if not council_name:
                print(
                    "[WARN] Skipping governorship without council name: "
                    f"{governorship.get('name', '<unknown>')}"
                )
                continue

            council_leader = str(
                ((governorship.get("council") or {}).get("leader") or {}).get("fullName") or ""
            ).strip()
            if council_name and council_name not in council_leaders:
                council_leaders[council_name] = council_leader

            rows_by_sheet[council_name].append(transform_governorship(governorship, current_week))

    for sheet_name, rows in rows_by_sheet.items():
        rows.sort(key=lambda item: _normalize(item.get("governorship")))
        print(f"[INFO] Prepared {len(rows)} row(s) for sheet '{sheet_name}'")

    return rows_by_sheet, council_leaders


def write_rows_to_sheet(ws: Worksheet, rows: List[Dict[str, Any]]) -> int:
    ws.cell(row=1, column=1, value=ws.title)
    for col_idx, header in enumerate(COLUMN_HEADERS, start=1):
        ws.cell(row=2, column=col_idx, value=header)

    current_row = 3
    for row in rows:
        ws.cell(row=current_row, column=1, value=row["governorship"])
        ws.cell(row=current_row, column=2, value=row["governor"])
        ws.cell(row=current_row, column=3, value=row["bacenta_count"])
        ws.cell(row=current_row, column=4, value=row["attendance"])
        ws.cell(row=current_row, column=5, value=row["income_ghs"])
        ws.cell(row=current_row, column=6, value=row["number_of_services"])
        ws.cell(row=current_row, column=7, value=row["services_by_bacentas"])
        ws.cell(row=current_row, column=8, value=row["services_not_held"])
        ws.cell(row=current_row, column=9, value=row["comment"])
        current_row += 1

    total_row = current_row
    ws.cell(row=total_row, column=1, value="TOTAL")

    formula_end = total_row - 1 if total_row > 3 else 2
    for col in ("C", "D", "E", "F", "H"):
        ws[f"{col}{total_row}"] = f"=SUM({col}3:{col}{formula_end})"

    ws[f"G{total_row}"] = f"=F{total_row}&\"/\"&C{total_row}"
    ws[f"I{total_row}"] = None

    return total_row


def create_summary_sheet(
    wb: Workbook,
    sheet_name_map: Dict[str, str],
    total_rows_by_sheet: Dict[str, int],
    council_leaders: Dict[str, str],
) -> None:
    summary_ws = wb.create_sheet(title="Summary", index=0)
    summary_ws.cell(row=1, column=1, value="SUMMARY")
    summary_ws.cell(row=2, column=1, value="Overseer")
    summary_ws.cell(row=2, column=2, value="Oversight Area")
    summary_ws.cell(row=2, column=3, value="Bacentas")
    summary_ws.cell(row=2, column=4, value="Att")
    summary_ws.cell(row=2, column=5, value="Income(GHS)")
    summary_ws.cell(row=2, column=6, value="No. Of Services")

    row_idx = 3
    for council_name in sorted(sheet_name_map.keys(), key=_normalize):
        detail_sheet = sheet_name_map[council_name]
        detail_total_row = total_rows_by_sheet.get(detail_sheet)
        if not detail_total_row:
            continue

        escaped_sheet = detail_sheet.replace("'", "''")
        summary_ws.cell(row=row_idx, column=1, value=council_leaders.get(council_name, ""))
        summary_ws.cell(row=row_idx, column=2, value=council_name)
        summary_ws.cell(row=row_idx, column=3, value=f"='{escaped_sheet}'!C{detail_total_row}")
        summary_ws.cell(row=row_idx, column=4, value=f"='{escaped_sheet}'!D{detail_total_row}")
        summary_ws.cell(row=row_idx, column=5, value=f"='{escaped_sheet}'!E{detail_total_row}")
        summary_ws.cell(row=row_idx, column=6, value=f"='{escaped_sheet}'!F{detail_total_row}")
        row_idx += 1


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Update Midweek data workbook from GraphQL stream data")
    parser.add_argument(
        "--input",
        default="Midweek_Data.xlsx",
        help="Legacy argument kept for compatibility; ignored when creating a new workbook",
    )
    parser.add_argument(
        "--output",
        default="Midweek_Data_updated.xlsx",
        help="Path to save updated workbook",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    output_path = Path(args.output)

    load_dotenv(Path(".env"))

    email = get_required_env("FLC_EMAIL")
    password = get_required_env("FLC_PASSWORD")

    stream_ids = [
        os.environ.get("STREAM_ID_1", DEFAULT_STREAM_1).strip(),
        os.environ.get("STREAM_ID_2", DEFAULT_STREAM_2).strip(),
        os.environ.get("STREAM_ID_3", DEFAULT_STREAM_3).strip(),
    ]
    stream_ids = [stream_id for stream_id in stream_ids if stream_id]
    if not stream_ids:
        raise RuntimeError("No stream IDs configured. Set STREAM_ID_1/2/3 environment variables.")

    current_week = date.today().isocalendar().week
    print(f"[INFO] Current ISO week: {current_week}")

    access_token = login_get_access_token(email, password)
    print("[INFO] Login successful; access token acquired")

    rows_by_sheet, council_leaders = collect_rows_by_sheet(stream_ids, access_token, current_week)

    workbook = Workbook()
    # Remove the auto-created default sheet so output has only generated tabs.
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    used_sheet_titles: set[str] = set()
    sheet_name_map: Dict[str, str] = {}
    total_rows_by_sheet: Dict[str, int] = {}
    written_counts: Dict[str, int] = {}

    for council_name in sorted(rows_by_sheet.keys(), key=_normalize):
        rows = rows_by_sheet.get(council_name, [])
        sheet_name = sanitize_sheet_title(council_name, used_sheet_titles)
        sheet_name_map[council_name] = sheet_name
        ws = workbook.create_sheet(title=sheet_name)
        total_row = write_rows_to_sheet(ws, rows)
        total_rows_by_sheet[sheet_name] = total_row
        written_counts[sheet_name] = len(rows)

    create_summary_sheet(workbook, sheet_name_map, total_rows_by_sheet, council_leaders)

    workbook.save(output_path)

    print("[INFO] Write complete")
    for sheet_name in sorted(written_counts.keys(), key=_normalize):
        print(f"[INFO] {sheet_name}: wrote {written_counts[sheet_name]} governorship row(s)")
    print(f"[INFO] Saved workbook: {output_path}")


if __name__ == "__main__":
    main()
