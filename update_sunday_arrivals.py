#!/usr/bin/env python3
"""Build Sunday_Arrivals.xlsx from Synago GraphQL council/governorship bussing data."""

from __future__ import annotations

import argparse
import os
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Dict, List, Tuple

import requests
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from update_midweek_data import (
    DEFAULT_STREAM_1,
    DEFAULT_STREAM_2,
    DEFAULT_STREAM_3,
    GRAPHQL_URL,
    _normalize,
    _safe_int,
    fetch_service_reports_for_stream,
    get_required_env,
    load_dotenv,
    login_get_access_token,
    sanitize_sheet_title,
)

# Mirrors the FLC_EMAIL_{n}/STREAM_ID_{n} configs: stream 1 is Colossians,
# stream 2 is the main Galatians councils, stream 3 is the Galatians councils
# that run a separate "Jesus Night" gathering on Sundays.
STREAM_GROUP_LABELS = {1: "Colossians", 2: "Galatians", 3: "Jesus Night"}

BUSSING_COUNCIL_QUERY = """
query BussingSubChurchesAtLevelCouncil(
  $id: ID!
  $startWeekKey: Int!
  $endWeekKey: Int!
  $targetLevel: String!
) {
  councils(where: { id: { eq: $id } }) {
    id
    name
    leader {
      fullName
    }
    subChurchesReportAtLevel(
      startWeekKey: $startWeekKey
      endWeekKey: $endWeekKey
      targetLevel: $targetLevel
    ) {
      churchId
      churchName
      week
      year
      bussingAttendance
      targetLeaderFirstName
      targetLeaderLastName
    }
  }
}
""".strip()


def fetch_bussing_for_council(
    council_id: str, access_token: str, start_week_key: int, end_week_key: int
) -> Tuple[str, List[Dict[str, Any]]]:
    payload = {
        "query": BUSSING_COUNCIL_QUERY,
        "variables": {
            "id": council_id,
            "startWeekKey": start_week_key,
            "endWeekKey": end_week_key,
            "targetLevel": "Governorship",
        },
    }
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {access_token}",
    }

    try:
        response = requests.post(GRAPHQL_URL, json=payload, headers=headers, timeout=45)
        response.raise_for_status()
    except requests.RequestException as exc:
        print(f"[WARN] Council {council_id}: bussing request failed: {exc}")
        return "", []

    body = response.json()
    if body.get("errors"):
        print(f"[WARN] Council {council_id}: bussing GraphQL errors: {body['errors']}")
        return "", []

    councils = body.get("data", {}).get("councils") or []
    if not councils:
        return "", []

    leader_name = str((councils[0].get("leader") or {}).get("fullName") or "").strip()
    entries = councils[0].get("subChurchesReportAtLevel") or []
    return leader_name, entries


def build_governorship_rows(
    entries: List[Dict[str, Any]], prev_week_key: int, curr_week_key: int
) -> List[Dict[str, Any]]:
    by_church: Dict[str, Dict[str, Any]] = {}
    for entry in entries:
        church_id = entry.get("churchId")
        if not church_id:
            continue
        bucket = by_church.setdefault(
            church_id,
            {
                "name": str(entry.get("churchName") or "").strip(),
                "governor": " ".join(
                    part
                    for part in (
                        entry.get("targetLeaderFirstName"),
                        entry.get("targetLeaderLastName"),
                    )
                    if part
                ).strip(),
                "prev": 0,
                "curr": 0,
            },
        )
        week_key = _safe_int(entry.get("year")) * 100 + _safe_int(entry.get("week"))
        if week_key == prev_week_key:
            bucket["prev"] = _safe_int(entry.get("bussingAttendance"))
        elif week_key == curr_week_key:
            bucket["curr"] = _safe_int(entry.get("bussingAttendance"))

    rows = list(by_church.values())
    rows.sort(key=lambda row: _normalize(row["name"]))
    return rows


def write_governorship_sheet(
    wb: Workbook, sheet_title: str, rows: List[Dict[str, Any]], prev_label: str, curr_label: str
) -> None:
    ws: Worksheet = wb.create_sheet(title=sheet_title)
    ws.cell(row=1, column=1, value="Area")
    ws.cell(row=1, column=2, value="Governor")
    ws.cell(row=1, column=3, value=prev_label)
    ws.cell(row=1, column=4, value=curr_label)

    row_idx = 2
    for row in rows:
        ws.cell(row=row_idx, column=1, value=row["name"])
        ws.cell(row=row_idx, column=2, value=row["governor"])
        ws.cell(row=row_idx, column=3, value=row["prev"])
        ws.cell(row=row_idx, column=4, value=row["curr"])
        row_idx += 1

    ws.cell(row=row_idx, column=1, value="TOTAL")
    ws.cell(row=row_idx, column=3, value=sum(row["prev"] for row in rows))
    ws.cell(row=row_idx, column=4, value=sum(row["curr"] for row in rows))


def collect_council_data(
    stream_configs: List[Tuple[int, str, str, str]], prev_week_key: int, curr_week_key: int
) -> List[Dict[str, Any]]:
    councils: List[Dict[str, Any]] = []

    for stream_index, stream_id, email, password in stream_configs:
        access_token = login_get_access_token(email, password)
        print(f"[INFO] Stream {stream_id}: login successful for {email}")

        council_lookup = fetch_service_reports_for_stream(
            stream_id, access_token, curr_week_key, target_level="Council"
        )
        print(f"[INFO] Stream {stream_id}: found {len(council_lookup)} council(s)")

        for council_id, council_info in council_lookup.items():
            council_name = str(council_info.get("churchName") or "").strip()
            if not council_name:
                continue

            leader_name, entries = fetch_bussing_for_council(
                council_id, access_token, prev_week_key, curr_week_key
            )
            rows = build_governorship_rows(entries, prev_week_key, curr_week_key)
            councils.append(
                {
                    "stream_index": stream_index,
                    "name": council_name,
                    "leader": leader_name,
                    "rows": rows,
                    "prev_total": sum(row["prev"] for row in rows),
                    "curr_total": sum(row["curr"] for row in rows),
                }
            )
            print(f"[INFO] Council '{council_name}': {len(rows)} governorship(s)")

    return councils


def write_rollup_sheet(
    wb: Workbook, councils: List[Dict[str, Any]], prev_label: str, curr_label: str
) -> None:
    ws: Worksheet = wb.create_sheet(title="Rollup", index=0)
    ws.cell(row=1, column=1, value="Council")
    ws.cell(row=1, column=2, value="Leader")
    ws.cell(row=1, column=3, value=prev_label)
    ws.cell(row=1, column=4, value=curr_label)

    councils_by_group: Dict[str, List[Dict[str, Any]]] = {}
    for council in councils:
        label = STREAM_GROUP_LABELS.get(council["stream_index"], f"Stream {council['stream_index']}")
        councils_by_group.setdefault(label, []).append(council)

    row_idx = 2
    grand_prev = 0
    grand_curr = 0
    for label in sorted(councils_by_group.keys()):
        group_prev = 0
        group_curr = 0
        for council in sorted(councils_by_group[label], key=lambda c: _normalize(c["name"])):
            ws.cell(row=row_idx, column=1, value=council["name"])
            ws.cell(row=row_idx, column=2, value=council["leader"])
            ws.cell(row=row_idx, column=3, value=council["prev_total"])
            ws.cell(row=row_idx, column=4, value=council["curr_total"])
            group_prev += council["prev_total"]
            group_curr += council["curr_total"]
            row_idx += 1

        ws.cell(row=row_idx, column=1, value=f"{label} Total")
        ws.cell(row=row_idx, column=3, value=group_prev)
        ws.cell(row=row_idx, column=4, value=group_curr)
        row_idx += 2
        grand_prev += group_prev
        grand_curr += group_curr

    ws.cell(row=row_idx, column=1, value="Grand Total")
    ws.cell(row=row_idx, column=3, value=grand_prev)
    ws.cell(row=row_idx, column=4, value=grand_curr)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build Sunday Arrivals workbook from bussing data")
    parser.add_argument("--output", default="Sunday_Arrivals_updated.xlsx", help="Path to save workbook")
    parser.add_argument(
        "--weeks-ago",
        type=int,
        default=0,
        help="Use the Sunday N weeks before today as the 'current' week (e.g. 1 for last Sunday)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_path = Path(args.output)

    load_dotenv(Path(".env"))

    stream_defaults = {1: DEFAULT_STREAM_1, 2: DEFAULT_STREAM_2, 3: DEFAULT_STREAM_3}
    stream_configs: List[Tuple[int, str, str, str]] = []
    for index, default_stream_id in stream_defaults.items():
        stream_id = os.environ.get(f"STREAM_ID_{index}", default_stream_id).strip()
        if not stream_id:
            continue
        email = get_required_env(f"FLC_EMAIL_{index}")
        password = get_required_env(f"FLC_PASSWORD_{index}")
        stream_configs.append((index, stream_id, email, password))

    if not stream_configs:
        raise RuntimeError("No stream IDs configured. Set STREAM_ID_1/2/3 environment variables.")

    target_date = date.today() - timedelta(weeks=args.weeks_ago)
    curr_iso = target_date.isocalendar()
    curr_week_key = curr_iso.year * 100 + curr_iso.week
    prev_iso = (target_date - timedelta(weeks=1)).isocalendar()
    prev_week_key = prev_iso.year * 100 + prev_iso.week

    curr_sunday = date.fromisocalendar(curr_iso.year, curr_iso.week, 7)
    prev_sunday = date.fromisocalendar(prev_iso.year, prev_iso.week, 7)
    curr_label = curr_sunday.strftime("%d/%m")
    prev_label = prev_sunday.strftime("%d/%m")
    print(f"[INFO] Comparing {prev_label} (weekKey={prev_week_key}) vs {curr_label} (weekKey={curr_week_key})")

    councils = collect_council_data(stream_configs, prev_week_key, curr_week_key)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    write_rollup_sheet(workbook, councils, prev_label, curr_label)

    used_sheet_titles: set[str] = set()
    for council in sorted(councils, key=lambda c: _normalize(c["name"])):
        sheet_name = sanitize_sheet_title(council["name"], used_sheet_titles)
        write_governorship_sheet(workbook, sheet_name, council["rows"], prev_label, curr_label)

    workbook.save(output_path)
    print(f"[INFO] Saved workbook: {output_path}")


if __name__ == "__main__":
    main()
